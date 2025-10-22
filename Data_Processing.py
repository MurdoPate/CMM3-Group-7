import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from scipy.interpolate import make_interp_spline
from openpyxl import load_workbook

# ----------------------------------------------------------
# Step 1: File paths
# ----------------------------------------------------------
input_file_path = r"C:\Users\sveng\OneDrive\Desktop\The University Of Edinburgh\Year 3\Computational methods and modelling 3\Group Project\Gait_Biomechanics_Dataset.csv"
output_file_path = r"C:\Users\sveng\OneDrive\Desktop\The University Of Edinburgh\Year 3\Computational methods and modelling 3\Group Project\Processed_Gait_Data.xlsx"

# ----------------------------------------------------------
# Step 2: Load dataset and filter for knee joint (2) and unbraced condition (1)
# ----------------------------------------------------------
data = pd.read_csv(input_file_path)
filtered_data = data[(data['joint'] == 2) & (data['condition'] == 1)].copy()

# ----------------------------------------------------------
# Step 3: Calculate mean and SD across gait cycle for each leg
# ----------------------------------------------------------
def calculate_leg_averages(data, parameter):
    """Compute mean and SD for each time point and leg."""
    stats = data.groupby(['time', 'leg'])[parameter].agg(['mean', 'std']).reset_index()
    left_leg = stats[stats['leg'] == 1].sort_values('time')
    right_leg = stats[stats['leg'] == 2].sort_values('time')
    return left_leg, right_leg, stats

angle_left, angle_right, angle_stats = calculate_leg_averages(filtered_data, 'angle')
velocity_left, velocity_right, velocity_stats = calculate_leg_averages(filtered_data, 'velocity')
accel_left, accel_right, accel_stats = calculate_leg_averages(filtered_data, 'acceleration')

# ----------------------------------------------------------
# Step 4: Create smoothed line plots (mean ± SD)
# ----------------------------------------------------------
def plot_leg_profile(leg_data, leg_label, parameter_name, y_label, color):
    """Plot mean ± SD profile with spline smoothing."""
    time_points = np.linspace(leg_data['time'].min(), leg_data['time'].max(), 200)
    mean_spline = make_interp_spline(leg_data['time'], leg_data['mean'], k=3)
    std_spline = make_interp_spline(leg_data['time'], leg_data['std'], k=3)

    mean_smooth = mean_spline(time_points)
    std_smooth = std_spline(time_points)

    plt.figure(figsize=(9,6))
    plt.plot(time_points, mean_smooth, color=color, linewidth=2, label=f'{leg_label} mean')
    plt.fill_between(time_points, mean_smooth - std_smooth, mean_smooth + std_smooth,
                     color=color, alpha=0.25, label='±1 SD')
    plt.xlabel('Gait cycle (%)')
    plt.ylabel(y_label)
    plt.title(f'{parameter_name} over Gait Cycle – {leg_label}')
    plt.legend()
    plt.grid(True, linestyle='--', alpha=0.7)
    plt.tight_layout()
    plt.show()

# ----------------------------------------------------------
# Step 5: Generate all plots
# ----------------------------------------------------------
# Left leg
plot_leg_profile(angle_left, 'Left leg', 'Knee Angle', 'Angle (radians)', 'blue')
plot_leg_profile(velocity_left, 'Left leg', 'Angular Velocity', 'Angular velocity (rad/s)', 'green')
plot_leg_profile(accel_left, 'Left leg', 'Angular Acceleration', 'Angular acceleration (rad/s²)', 'purple')

# Right leg
plot_leg_profile(angle_right, 'Right leg', 'Knee Angle', 'Angle (radians)', 'red')
plot_leg_profile(velocity_right, 'Right leg', 'Angular Velocity', 'Angular velocity (rad/s)', 'orange')
plot_leg_profile(accel_right, 'Right leg', 'Angular Acceleration', 'Angular acceleration (rad/s²)', 'brown')

# ----------------------------------------------------------
# Step 6: Create summary tables for Excel
# ----------------------------------------------------------
def create_summary_table(stats, parameter):
    """Pivot mean and SD values into a tidy Excel-friendly format."""
    stats['leg'] = stats['leg'].replace({1: 'Left', 2: 'Right'})
    summary = stats.pivot(index='time', columns='leg', values=['mean','std'])
    summary.columns = [f'{parameter}_{col[0]}_{col[1]}' for col in summary.columns]
    summary.reset_index(inplace=True)
    summary.rename(columns={'time': 'Gait cycle (%)'}, inplace=True)
    return summary

angle_summary = create_summary_table(angle_stats, 'Angle')
velocity_summary = create_summary_table(velocity_stats, 'Velocity')
accel_summary = create_summary_table(accel_stats, 'Acceleration')

# ----------------------------------------------------------
# Step 7: Export processed data and summaries to Excel
# ----------------------------------------------------------
with pd.ExcelWriter(output_file_path, engine='openpyxl') as writer:
    filtered_data.to_excel(writer, sheet_name='Filtered Raw Data', index=False)
    angle_summary.to_excel(writer, sheet_name='Knee Angle (Mean_SD)', index=False)
    velocity_summary.to_excel(writer, sheet_name='Angular Velocity (Mean_SD)', index=False)
    accel_summary.to_excel(writer, sheet_name='Angular Acceleration (Mean_SD)', index=False)

# ----------------------------------------------------------
# Step 8: Adjust column widths for better readability
# ----------------------------------------------------------
wb = load_workbook(output_file_path)
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    for column_cells in ws.columns:
        max_length = max((len(str(cell.value)) for cell in column_cells if cell.value is not None), default=10)
        ws.column_dimensions[column_cells[0].column_letter].width = max_length + 3
wb.save(output_file_path)
wb.close()